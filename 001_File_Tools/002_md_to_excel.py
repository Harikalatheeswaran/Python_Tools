#!/usr/bin/env python3
"""
Markdown-to-Excel Converter

- Reads a Markdown string or file containing multiple tables.
- Creates an Excel workbook with:
  1) A "Legend & TOC" sheet explaining colors, emoji meanings, and links to each table sheet.
  2) Each Markdown table placed on its own sheet with formatting:
     - Header styling, borders, alternate row shading
     - Auto column widths, text wrap, frozen header row
     - Emoji-based cell coloring (✅, ❌, ⚠️)

Usage:
  python md_to_excel.py -i benchmark.md -o benchmark.xlsx

If -i is omitted, script reads Markdown from STDIN.
Requires: pip install openpyxl
"""

import re
import sys
import argparse
from typing import List, Dict, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color
from openpyxl.utils import get_column_letter

# ---------------------------
# Parsing Markdown Tables
# ---------------------------

def clean_text(s: str) -> str:
    if s is None:
        return ""
    # Replace HTML breaks with newlines
    s = s.replace("<br><br>", "\n\n").replace("<br />", "\n").replace("<br>", "\n")
    # Strip Markdown bold/italic markers
    s = s.replace("**", "").replace("__", "").replace("*", "")
    # Trim whitespace
    return s.strip()

def is_table_separator(line: str) -> bool:
    # A typical markdown separator: |---|:---:|---|
    return bool(re.match(r'^\s*\|?(\s*:?-+:?\s*\|)+\s*:?-+:?\s*\|?\s*$', line))

def split_md_row(line: str) -> List[str]:
    # Split by pipe, ignore the first/last empty if line starts/ends with '|'
    parts = [clean_text(p) for p in line.strip().strip('|').split('|')]
    return parts

def extract_heading_context(lines: List[str], idx: int) -> str:
    # Walk backwards to find the most recent heading (#, ##, ###) to name the sheet
    for j in range(idx - 1, -1, -1):
        m = re.match(r'^\s{0,3}(#{1,6})\s+(.*)$', lines[j])
        if m:
            return clean_text(m.group(2))[:60]  # limit length for sheet name base
    return "Table"

def parse_markdown_tables(md_text: str) -> List[Dict]:
    """
    Returns a list of dicts: {'title': str, 'headers': List[str], 'rows': List[List[str]]}
    """
    lines = md_text.splitlines()
    tables = []
    i = 0
    while i < len(lines):
        line = lines[i]
        # Detect table header row (must contain at least 2 pipes and next line must be separator)
        if '|' in line and line.count('|') >= 2 and i + 1 < len(lines) and is_table_separator(lines[i + 1]):
            headers = split_md_row(line)
            i += 2  # skip header and separator
            rows = []
            while i < len(lines) and '|' in lines[i] and not lines[i].strip().startswith('---'):
                row_parts = split_md_row(lines[i])
                # pad/truncate to header length to avoid ragged rows
                if len(row_parts) < len(headers):
                    row_parts += [""] * (len(headers) - len(row_parts))
                elif len(row_parts) > len(headers):
                    row_parts = row_parts[:len(headers)]
                rows.append(row_parts)
                i += 1
            title = extract_heading_context(lines, i)
            tables.append({'title': title, 'headers': headers, 'rows': rows})
            continue
        i += 1
    return tables

# ---------------------------
# Excel Writing & Styling
# ---------------------------

HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_ROW_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # Light blue
BORDER_THIN = Border(left=Side(style='thin', color='888888'),
                     right=Side(style='thin', color='888888'),
                     top=Side(style='thin', color='888888'),
                     bottom=Side(style='thin', color='888888'))
WRAP = Alignment(wrap_text=True, vertical="top")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOC_HEADER_FILL = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")

def sanitize_sheet_name(name: str, existing: set) -> str:
    # Excel sheet name rules: max 31 chars; exclude : \ / ? * [ ]
    cleaned = re.sub(r'[:\\/\?\*\[\]]', '_', name).strip()
    if len(cleaned) == 0:
        cleaned = "Sheet"
    base = cleaned[:31]
    # Ensure uniqueness
    sheet_name = base
    counter = 2
    while sheet_name in existing or len(sheet_name) == 0:
        suffix = f" ({counter})"
        sheet_name = (base[:31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        counter += 1
    existing.add(sheet_name)
    return sheet_name

def autosize_columns(ws):
    max_col = ws.max_column
    widths = [0] * max_col
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row):
            length = len(str(val)) if val is not None else 0
            if length > widths[idx]:
                widths[idx] = length
    for idx, w in enumerate(widths, start=1):
        col_letter = get_column_letter(idx)
        # heuristic width scaling; cap at 80 chars
        ws.column_dimensions[col_letter].width = min(max(w * 0.9, 15), 80)

def apply_table_styles(ws):
    # Header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER_THIN
    # Body rows
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER_THIN
            # Alternate row shading
            if r % 2 == 0:
                cell.fill = ALT_ROW_FILL
            # Emoji-based coloring
            val = str(cell.value) if cell.value is not None else ""
            if "✅" in val:
                cell.fill = GREEN_FILL
            elif "❌" in val:
                cell.fill = RED_FILL
            elif "⚠️" in val or "⚠" in val:
                cell.fill = YELLOW_FILL
    # Freeze header row and enable filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def write_table_sheet(wb: Workbook, title: str, headers: List[str], rows: List[List[str]]) -> str:
    sheet_name = sanitize_sheet_name(title, set(wb.sheetnames))
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for r in rows:
        ws.append(r)
    apply_table_styles(ws)
    autosize_columns(ws)
    # Increase row height for header
    ws.row_dimensions[1].height = 28
    return sheet_name

def create_legend_sheet(wb: Workbook, table_summaries: List[Tuple[str, int, int]]):
    ws = wb.active
    ws.title = "Legend & TOC"

    # Title
    ws["A1"] = "AI Model Benchmark Analysis: AES Portfolio RAG Performance"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # Legend headers
    ws["A3"] = "Legend"
    ws["A3"].font = Font(bold=True, size=12)
    ws["A3"].fill = TOC_HEADER_FILL
    ws["A3"].font = Font(color="FFFFFF", bold=True)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)

    # Legend content
    legend_rows = [
        ("Header Fill", "Blue background, white bold text", ""),
        ("Alternate Rows", "Light blue shading", ""),
        ("✅", "Correct / Accurate", ""),
        ("❌", "Incorrect / Mismatch", ""),
        ("⚠️", "Partial / Ambiguous", ""),
    ]
    ws.append(["Symbol/Style", "Meaning", "Sample"])
    for cell in ws[4]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
        cell.alignment = Alignment(vertical="center")

    for label, meaning, _ in legend_rows:
        ws.append([label, meaning, ""])
        r = ws.max_row
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = BORDER_THIN
            ws.cell(row=r, column=c).alignment = WRAP
        # Apply corresponding fills
        if label == "Header Fill":
            ws.cell(row=r, column=1).fill = HEADER_FILL
            ws.cell(row=r, column=1).font = HEADER_FONT
        elif label == "Alternate Rows":
            ws.cell(row=r, column=1).fill = ALT_ROW_FILL
        elif label == "✅":
            ws.cell(row=r, column=1).fill = GREEN_FILL
        elif label == "❌":
            ws.cell(row=r, column=1).fill = RED_FILL
        elif "⚠" in label:
            ws.cell(row=r, column=1).fill = YELLOW_FILL

    # TOC Header
    ws.append([])
    start_toc = ws.max_row + 1
    ws.cell(row=start_toc, column=1, value="Tables (Click to open)").fill = TOC_HEADER_FILL
    ws.cell(row=start_toc, column=1).font = Font(color="FFFFFF", bold=True)
    ws.merge_cells(start_row=start_toc, start_column=1, end_row=start_toc, end_column=4)

    # TOC Table
    ws.append(["Sheet Name", "Rows", "Columns", "Link"])
    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN

    for name, nrows, ncols in table_summaries:
        ws.append([name, nrows, ncols, f"Go to {name}"])
        r = ws.max_row
        ws.cell(row=r, column=4).hyperlink = f"#{name}!A1"
        ws.cell(row=r, column=4).font = Font(color="0563C1", underline="single")
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(vertical="center")

    autosize_columns(ws)
    ws.freeze_panes = "A6"

# ---------------------------
# Conversion Orchestrator
# ---------------------------

def convert_markdown_to_excel(md_text: str, out_path: str = "benchmark.xlsx"):
    tables = parse_markdown_tables(md_text)
    if not tables:
        raise ValueError("No markdown tables were found in the input.")

    wb = Workbook()
    # Create table sheets and collect summaries
    table_summaries = []
    sheet_names = []
    # Reserve first sheet for Legend
    ws_legend = wb.active
    ws_legend.title = "Legend & TOC"

    for t in tables:
        name = write_table_sheet(wb, t['title'], t['headers'], t['rows'])
        sheet_names.append(name)
        # Determine dimensions (excluding header)
        nrows = len(t['rows'])
        ncols = len(t['headers'])
        table_summaries.append((name, nrows, ncols))

    # Move Legend to first position
    wb.move_sheet("Legend & TOC", offset=-len(wb.sheetnames))
    # Populate Legend
    create_legend_sheet(wb, table_summaries)

    wb.save(out_path)
    print(f"Excel file written to: {out_path}")

# ---------------------------
# CLI
# ---------------------------

def main():
    parser = argparse.ArgumentParser(description="Convert Markdown tables to a formatted Excel workbook.")
    parser.add_argument("-i", "--input", help="Path to Markdown file. If omitted, read from STDIN.", default=None)
    parser.add_argument("-o", "--output", help="Output Excel file path", default="benchmark.xlsx")
    args = parser.parse_args()

    if args.input:
        with open(args.input, "r", encoding="utf-8") as f:
            md_text = f.read()
    else:
        md_text = sys.stdin.read()

    convert_markdown_to_excel(md_text, args.output)

if __name__ == "__main__":
    main()